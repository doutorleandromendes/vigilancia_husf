"""
parse_alcool_uti.py — extrai consumo de álcool gel das UTIs (A/B e C)
da aba 'Dados UTI' da Planilha Geral de Controles.

Fonte primária: contagem de bags de 500ml reposta pelo SHL no setor.
Formato (texto livre, col 1): 'UTI A e B: 29 + 12 UTI C: 23 + 13 Sabonete ...'
  → soma dos números após 'UTI A e B' = bags A/B; após 'UTI C' = bags C
  → ml = bags × 500
Pac-dia por UTI vem do CVE Plan2 (não da Planilha de Controles).

HM (ml/pac-dia) = (bags × 500) / pac-dia
"""
import re
from openpyxl import load_workbook

ABBR = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
MESES = ['JANEIRO','FEVEREIRO','MARÇO','ABRIL','MAIO','JUNHO',
         'JULHO','AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']
DATA_START = {2025: 4, 2026: 4}  # linha de Janeiro na aba Dados UTI
BAG_ML = 500

def _norm(s):
    return str(s).strip().upper() if s else ''

def _bags(txt, label):
    """Soma bags após `label` até o próximo marcador (UTI/Sabonete)."""
    if not txt: return None
    # captura tudo após label até 'UTI' (outra) ou 'Sabonete'
    pat = label + r'\s*:?\s*([\d\s\+]+?)(?=UTI|Sabonete|$)'
    m = re.search(pat, txt, re.IGNORECASE)
    if not m: return None
    nums = re.findall(r'\d+', m.group(1))
    return sum(int(n) for n in nums) if nums else None

def parse_alcool_uti(controles_path, ano):
    """
    Retorna {'utiAB': {mes_label: bags_ml}, 'utic': {mes_label: bags_ml}}
    (ainda sem dividir por pac-dia — isso vem depois com o CVE Plan2).
    """
    wb = load_workbook(controles_path, read_only=True, data_only=True)
    ws = wb['Dados UTI']
    rows = list(ws.iter_rows(values_only=True))
    out = {'utiAB': {}, 'utic': {}}
    for mi in range(12):
        # localiza a linha do mês
        ri = None
        for i, r in enumerate(rows[:20]):
            if r and _norm(r[0]) == MESES[mi]:
                ri = i; break
        if ri is None: continue
        txt = rows[ri][1]
        if not txt: continue
        ab = _bags(txt, r'UTI A e B')
        c  = _bags(txt, r'UTI\s+C')
        label = f"{ABBR[mi]}/{str(ano)[-2:]}"
        if ab: out['utiAB'][label] = ab * BAG_ML
        if c:  out['utic'][label]  = c * BAG_ML
    return out

if __name__ == '__main__':
    import sys, json
    for fname, ano in [(sys.argv[1], int(sys.argv[2]))]:
        r = parse_alcool_uti(fname, ano)
        print(json.dumps(r, indent=2, ensure_ascii=False))
