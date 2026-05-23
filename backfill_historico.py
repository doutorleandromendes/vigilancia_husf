#!/usr/bin/env python3
# =============================================================================
# backfill_historico.py  ·  v1.0  ·  Mai 2026
# Lê arquivos XLSX históricos de DOT e MDR (2024-2025) e gera
# data_mensal_hist.json — série mensal granular para análise estatística.
#
# Dr. Leandro Mendes · SCIH-CCIH · HUSF · Bragança Paulista
# =============================================================================
#
# Dependência: openpyxl
#   pip install openpyxl   OU   pip3 install openpyxl
#
# USO:
#   python3 backfill_historico.py
#
# O script procura arquivos nos diretórios configurados abaixo.
# Ajuste as variáveis na seção CONFIGURAÇÃO conforme necessário.
# =============================================================================

import os
import json
import re
import glob
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
    raise

# =============================================================================
# ── CONFIGURAÇÃO — ajuste estes caminhos ─────────────────────────────────────
# =============================================================================

# Diretório onde estão os arquivos DOT de anos anteriores.
# Nomes esperados: DOTs_2024.xlsx, DOTs_2025.xlsx
# (qualquer arquivo com "DOT" no nome e extensão .xlsx será tentado)
DOTS_DIR = "."

# Diretório onde estão os arquivos de censo MDR mensais de anos anteriores.
# Nomes esperados: MDR_Jan2024.xlsx, MDR_Fev2024.xlsx ... ou similares.
# (qualquer arquivo com "MDR" ou "censo" no nome será tentado)
MDR_DIR = "."

# Arquivo de saída
OUTPUT_FILE = "data_mensal_hist.json"

# Ano mínimo para incluir na série histórica
ANO_MINIMO = 2024

# =============================================================================
# ── CONSTANTES (replicam exatamente a lógica de entrada_dados.html) ──────────
# =============================================================================

MESES_PT = ["jan", "fev", "mar", "abr", "mai", "jun",
            "jul", "ago", "set", "out", "nov", "dez"]

MESES_PT_UPPER = [m.upper() for m in MESES_PT]

MESES_PT_ABBR = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                 "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

MESES_EN = ["January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"]

MESES_EN_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Mapeamento de rótulo do Excel → índice 0-11
def parse_mes_label(label):
    """Converte qualquer rótulo de mês em índice 0-11. Retorna None se falhar."""
    if label is None:
        return None
    s = str(label).strip().upper()
    # Tentativa 1: português maiúsculo
    for i, m in enumerate(MESES_PT_UPPER):
        if s.startswith(m[:3]):
            return i
    # Tentativa 2: inglês
    for i, m in enumerate(MESES_EN):
        if s.startswith(m[:3].upper()):
            return i
    # Tentativa 3: número
    try:
        n = int(float(s))
        if 1 <= n <= 12:
            return n - 1
    except (ValueError, TypeError):
        pass
    return None

def mes_label(mes_idx, ano):
    """Gera o rótulo padrão do pipeline: 'jan/24', 'fev/25', etc."""
    return f"{MESES_PT[mes_idx]}/{str(ano)[2:]}"

def safe_float(v):
    """Converte valor de célula em float, retorna None se inválido."""
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", ".").strip())
        return None if (f != f) else round(f, 4)  # NaN check
    except (ValueError, TypeError):
        return None

# =============================================================================
# ── PARSING DE DOT ────────────────────────────────────────────────────────────
#
# Estrutura da aba Tot_XXX (espelhada de entrada_dados.html):
#   row[0]: cabeçalho ATB  (ignorar)
#   row[1]: cabeçalho unid (ignorar)
#   row[2..13]: JAN..DEZ
#     col[0] = nome do mês
#     col[1] = UTI A/B
#     col[2] = UTIC
#     col[3] = Clínica Médica
#     col[4] = Clínica Cirúrgica
#     col[5] = Apartamentos
#     col[6] = EPM
#   row[14+]: contagens absolutas (ignorar)
# =============================================================================

DOTS_TABS = [
    {"sheet": "Tot_CTO",  "key": "cef"},
    {"sheet": "Tot_PPTZ", "key": "pip"},
    {"sheet": "Tot_CBP",  "key": "cbp"},
    {"sheet": "Tot_GPP",  "key": "gpp"},
    {"sheet": "Tot_PB",   "key": "pb"},
]

DOTS_UNIT_COLS = [
    {"field": "utiab", "col": 1},
    {"field": "utic",  "col": 2},
    {"field": "clin",  "col": 3},
    {"field": "cir",   "col": 4},
    {"field": "apto",  "col": 5},
    {"field": "epm",   "col": 6},
]

def parse_dots_xlsx(filepath, ano):
    """
    Lê um arquivo DOT XLSX e retorna dict:
    {
      "cef": {"utiab": [{"p":"jan/24","v":12.3}, ...], "utic": [...], ...},
      "pip": {...},
      ...
    }
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ✗ Erro ao abrir {filepath}: {e}")
        return {}

    result = {}
    for tab in DOTS_TABS:
        key = tab["key"]
        sheet_name = tab["sheet"]

        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ {Path(filepath).name}: aba '{sheet_name}' não encontrada")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))

        if key not in result:
            result[key] = {u["field"]: [] for u in DOTS_UNIT_COLS}

        # Linhas 3..14 (1-indexed) = rows[2..13] (0-indexed)
        for row_idx in range(2, min(14, len(rows))):
            row = rows[row_idx]
            mes_idx = parse_mes_label(row[0] if row else None)
            if mes_idx is None:
                continue

            label = mes_label(mes_idx, ano)

            for uc in DOTS_UNIT_COLS:
                v = safe_float(row[uc["col"]]) if len(row) > uc["col"] else None
                if v is not None:
                    result[key][uc["field"]].append({"p": label, "v": v})

    wb.close()
    return result

# =============================================================================
# ── PARSING DE MDR MENSAL ─────────────────────────────────────────────────────
#
# Estrutura (replicada de parseMDRCenso em entrada_dados.html):
# Uma aba por mês (nome: mês em inglês, português abreviado, ou número).
# Linha "Total" seguida de linha de densidades "Di".
#
# UTI A/B  (cols 1–9):
#   col[1]=ESBL_cum  [2]=ESBL_novo  [3]=Acin_cum   [4]=Acin_novo
#   col[5]=KPC_cum   [6]=KPC_novo   [7]=Pseu_cum   [8]=Pseu_novo  [9]=pd
#
# UTI C    (cols 13–21, offset=12 sobre UTI A/B):
#   mesma estrutura, offset uc_off=12
#
# UTI Neo  (cols 25–33, offset=24):
#   mesma estrutura, offset neo_off=24
# =============================================================================

def find_mes_sheet(wb, mes_idx):
    """Encontra a aba do mês pelo índice 0-11. Retorna sheet ou None."""
    candidates = [
        MESES_EN[mes_idx],           # "January"
        MESES_PT_ABBR[mes_idx],      # "Jan"
        MESES_EN_ABBR[mes_idx],      # "Jan" (inglês)
        str(mes_idx + 1),            # "1"
        f"{mes_idx+1:02d}",          # "01"
    ]
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    # Busca parcial case-insensitive
    for sn in wb.sheetnames:
        for c in candidates:
            if c.lower() in sn.lower():
                return wb[sn]
    return None

def parse_mdr_sheet(ws, mes_idx, ano):
    """
    Extrai dados MDR de uma aba mensal. Retorna dict ou None se falhar.
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    def safe_n(row, col):
        if row is None or col >= len(row):
            return 0
        v = safe_float(row[col])
        return v if v is not None else 0

    # Encontrar a linha "Total" (última ocorrência de "total" em col[0])
    total_row_idx = -1
    for r_idx in range(len(rows) - 1, -1, -1):
        cell = rows[r_idx][0] if rows[r_idx] else None
        if cell is not None and str(cell).strip().lower() == "total":
            total_row_idx = r_idx
            break

    if total_row_idx < 0:
        # Fallback: linha 35 (índice 34) como em entrada_dados.html
        total_row_idx = min(34, len(rows) - 2)

    if total_row_idx >= len(rows):
        return None

    tot = rows[total_row_idx] if total_row_idx < len(rows) else []
    di  = rows[total_row_idx + 1] if (total_row_idx + 1) < len(rows) else []

    tot = list(tot) if tot else []
    di  = list(di)  if di  else []

    def rate(n, pd):
        return round(n / pd * 1000, 2) if pd > 0 else 0.0

    def extract_unit(offset):
        pd       = safe_n(tot, offset + 9)
        esbl_n   = safe_n(tot, offset + 2)
        acin_n   = safe_n(tot, offset + 4)
        kpc_n    = safe_n(tot, offset + 6)
        pseu_n   = safe_n(tot, offset + 8)
        esbl_r   = safe_n(di,  offset + 2) or rate(esbl_n, pd)
        acin_r   = safe_n(di,  offset + 4) or rate(acin_n, pd)
        kpc_r    = safe_n(di,  offset + 6) or rate(kpc_n,  pd)
        pseu_r   = safe_n(di,  offset + 8) or rate(pseu_n, pd)
        return {
            "pd": int(pd), "esbl_n": int(esbl_n), "acin_n": int(acin_n),
            "kpc_n": int(kpc_n), "pseu_n": int(pseu_n),
            "esbl_r": esbl_r, "acin_r": acin_r,
            "kpc_r": kpc_r, "pseu_r": pseu_r,
        }

    return {
        "label": mes_label(mes_idx, ano),
        "ab":  extract_unit(0),
        "uc":  extract_unit(12),
        "neo": extract_unit(24),
    }

def parse_mdr_xlsx(filepath, ano):
    """
    Lê um arquivo MDR e retorna lista de dicts mensais:
    [{"label":"jan/24", "ab":{...}, "uc":{...}, "neo":{...}}, ...]

    Aceita dois formatos:
      1. Um arquivo por ano com uma aba por mês (mais comum)
      2. Um arquivo por mês (aba única, nome do mês ou genérico)
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ✗ Erro ao abrir {filepath}: {e}")
        return []

    results = []

    # Tenta encontrar abas para cada mês
    for mes_idx in range(12):
        ws = find_mes_sheet(wb, mes_idx)
        if ws is None:
            continue
        parsed = parse_mdr_sheet(ws, mes_idx, ano)
        if parsed and parsed["ab"]["pd"] > 0:
            results.append(parsed)

    # Se nenhuma aba por mês encontrada, tenta aba única (arquivo por mês)
    if not results and len(wb.sheetnames) >= 1:
        # Tenta inferir o mês pelo nome do arquivo
        fname = Path(filepath).stem.upper()
        mes_idx = None
        for i, m in enumerate(MESES_PT_UPPER + MESES_EN):
            if m[:3] in fname:
                mes_idx = i % 12
                break
        if mes_idx is not None:
            ws = wb.active
            parsed = parse_mdr_sheet(ws, mes_idx, ano)
            if parsed and parsed["ab"]["pd"] > 0:
                results.append(parsed)

    wb.close()
    return results

# =============================================================================
# ── DESCOBERTA AUTOMÁTICA DE ARQUIVOS ─────────────────────────────────────────
# =============================================================================

def find_dot_files(directory):
    """Encontra arquivos DOT XLSX no diretório."""
    files = []
    for pattern in ["*DOT*.xlsx", "*dot*.xlsx", "*DOT*.XLSX"]:
        files.extend(glob.glob(os.path.join(directory, pattern)))
    return sorted(set(files))

def find_mdr_files(directory):
    """Encontra arquivos MDR/censo XLSX no diretório."""
    files = []
    for pattern in ["*MDR*.xlsx", "*mdr*.xlsx", "*censo*MDR*.xlsx",
                     "*Censo*MDR*.xlsx", "*CENSO*MDR*.xlsx"]:
        files.extend(glob.glob(os.path.join(directory, pattern)))
    return sorted(set(files))

def infer_ano(filepath):
    """Tenta inferir o ano do nome do arquivo. Retorna None se não encontrar."""
    fname = Path(filepath).stem
    match = re.search(r'(20\d{2})', fname)
    if match:
        return int(match.group(1))
    return None

# =============================================================================
# ── MERGE DE SÉRIES ───────────────────────────────────────────────────────────
# =============================================================================

def merge_dot_series(all_dots):
    """
    Combina dicts de DOT de vários anos em séries ordenadas cronologicamente.
    all_dots = lista de dicts (um por arquivo/ano), cada um no formato:
      {"cef": {"utiab": [{"p":"jan/24","v":x}, ...], ...}, ...}
    Retorna um único dict com as séries concatenadas e ordenadas.
    """
    merged = {}
    for atb_dict in all_dots:
        for atb_key, unit_dict in atb_dict.items():
            if atb_key not in merged:
                merged[atb_key] = {}
            for unit_key, pts in unit_dict.items():
                if unit_key not in merged[atb_key]:
                    merged[atb_key][unit_key] = {}
                for pt in pts:
                    # Usa p como chave para deduplicar
                    merged[atb_key][unit_key][pt["p"]] = pt["v"]

    # Converte para lista ordenada
    def sort_key(label):
        m, a = label.split("/")
        return int(a) * 100 + MESES_PT.index(m)

    result = {}
    for atb_key, unit_dict in merged.items():
        result[atb_key] = {}
        for unit_key, pts_dict in unit_dict.items():
            sorted_labels = sorted(pts_dict.keys(), key=sort_key)
            result[atb_key][unit_key] = [
                {"p": lbl, "v": pts_dict[lbl]} for lbl in sorted_labels
            ]
    return result

def merge_mdr_series(all_mdr_months):
    """
    Combina lista de meses MDR em séries por unidade/organismo.
    Retorna estrutura compatível com mdrMensal do data_iras.json.
    """
    def sort_key(label):
        m, a = label.split("/")
        return int(a) * 100 + MESES_PT.index(m)

    # Deduplicar por label
    by_label = {}
    for month in all_mdr_months:
        by_label[month["label"]] = month

    sorted_labels = sorted(by_label.keys(), key=sort_key)

    utiAB = {"esbl": [], "kpc": [], "acin": [], "pseu": [], "pd": [], "counts": {
        "esbl": [], "kpc": [], "acin": [], "pseu": []}}
    utic  = {"esbl": [], "kpc": [], "acin": [], "pseu": [], "pd": [], "counts": {
        "esbl": [], "kpc": [], "acin": [], "pseu": []}}

    for lbl in sorted_labels:
        m = by_label[lbl]

        ab = m["ab"]
        utiAB["esbl"].append({"p": lbl, "v": ab["esbl_r"], "pd": ab["pd"]})
        utiAB["kpc"].append( {"p": lbl, "v": ab["kpc_r"],  "pd": ab["pd"]})
        utiAB["acin"].append({"p": lbl, "v": ab["acin_r"], "pd": ab["pd"]})
        utiAB["pseu"].append({"p": lbl, "v": ab["pseu_r"], "pd": ab["pd"]})
        utiAB["pd"].append(  {"p": lbl, "v": ab["pd"]})
        utiAB["counts"]["esbl"].append(ab["esbl_n"])
        utiAB["counts"]["kpc"].append( ab["kpc_n"])
        utiAB["counts"]["acin"].append(ab["acin_n"])
        utiAB["counts"]["pseu"].append(ab["pseu_n"])

        uc = m["uc"]
        utic["esbl"].append({"p": lbl, "v": uc["esbl_r"], "pd": uc["pd"]})
        utic["kpc"].append( {"p": lbl, "v": uc["kpc_r"],  "pd": uc["pd"]})
        utic["acin"].append({"p": lbl, "v": uc["acin_r"], "pd": uc["pd"]})
        utic["pseu"].append({"p": lbl, "v": uc["pseu_r"], "pd": uc["pd"]})
        utic["pd"].append(  {"p": lbl, "v": uc["pd"]})
        utic["counts"]["esbl"].append(uc["esbl_n"])
        utic["counts"]["kpc"].append( uc["kpc_n"])
        utic["counts"]["acin"].append(uc["acin_n"])
        utic["counts"]["pseu"].append(uc["pseu_n"])

    return {"utiAB": utiAB, "utic": utic}

# =============================================================================
# ── MAIN ──────────────────────────────────────────────────────────────────────
# =============================================================================

def main():
    print("=" * 65)
    print("  backfill_historico.py  ·  HUSF / SCIH-CCIH")
    print("=" * 65)

    # ── DOTs ──────────────────────────────────────────────────────────────────
    print("\n[1/2] Processando arquivos DOT...")
    dot_files = find_dot_files(DOTS_DIR)

    if not dot_files:
        print(f"  ⚠ Nenhum arquivo DOT encontrado em '{DOTS_DIR}'")
        print("    Ajuste DOTS_DIR no topo do script.")
    else:
        print(f"  Encontrados: {[Path(f).name for f in dot_files]}")

    all_dots = []
    for fpath in dot_files:
        ano = infer_ano(fpath)
        if ano is None:
            print(f"  ⚠ Não foi possível inferir o ano de '{Path(fpath).name}' — pulando")
            continue
        if ano < ANO_MINIMO:
            print(f"  → {Path(fpath).name}: ano {ano} < {ANO_MINIMO}, pulando")
            continue
        print(f"  → {Path(fpath).name} (ano {ano})...")
        parsed = parse_dots_xlsx(fpath, ano)
        if parsed:
            # Contar pontos extraídos
            n_pts = sum(len(v) for atb in parsed.values() for v in atb.values())
            print(f"    ✓ {n_pts} pontos mensais extraídos")
            all_dots.append(parsed)
        else:
            print(f"    ✗ Nenhum dado extraído")

    dots_merged = merge_dot_series(all_dots) if all_dots else {}

    # ── MDR ───────────────────────────────────────────────────────────────────
    print("\n[2/2] Processando arquivos MDR...")
    mdr_files = find_mdr_files(MDR_DIR)

    if not mdr_files:
        print(f"  ⚠ Nenhum arquivo MDR encontrado em '{MDR_DIR}'")
        print("    Ajuste MDR_DIR no topo do script.")
    else:
        print(f"  Encontrados: {[Path(f).name for f in mdr_files]}")

    all_mdr_months = []
    for fpath in mdr_files:
        ano = infer_ano(fpath)
        if ano is None:
            print(f"  ⚠ Ano não inferido de '{Path(fpath).name}' — tentando mesmo assim...")
            # Tenta com ano atual como fallback (não ideal)
            ano = 2024
        if ano < ANO_MINIMO:
            print(f"  → {Path(fpath).name}: ano {ano} < {ANO_MINIMO}, pulando")
            continue
        print(f"  → {Path(fpath).name} (ano {ano})...")
        months = parse_mdr_xlsx(fpath, ano)
        if months:
            print(f"    ✓ {len(months)} meses extraídos: {[m['label'] for m in months]}")
            all_mdr_months.extend(months)
        else:
            print(f"    ✗ Nenhum mês extraído — verifique o formato da aba")

    mdr_merged = merge_mdr_series(all_mdr_months) if all_mdr_months else {}

    # ── Sumário ───────────────────────────────────────────────────────────────
    print("\n── Sumário ──────────────────────────────────────────────")
    if dots_merged:
        for atb_key in dots_merged:
            pts_count = {u: len(v) for u, v in dots_merged[atb_key].items()}
            primeiro = next(iter(dots_merged[atb_key].values()))[0]["p"] if any(dots_merged[atb_key].values()) else "?"
            ultimo   = next(iter(dots_merged[atb_key].values()))[-1]["p"] if any(dots_merged[atb_key].values()) else "?"
            print(f"  DOT {atb_key:5s}: {pts_count} pts  ({primeiro} → {ultimo})")
    else:
        print("  DOTs: nenhum dado processado")

    if mdr_merged:
        ab_labels = [pt["p"] for pt in mdr_merged["utiAB"]["pd"]]
        print(f"  MDR UTI A/B: {len(ab_labels)} meses  ({ab_labels[0] if ab_labels else '?'} → {ab_labels[-1] if ab_labels else '?'})")
        uc_labels = [pt["p"] for pt in mdr_merged["utic"]["pd"]]
        print(f"  MDR UTI C:   {len(uc_labels)} meses  ({uc_labels[0] if uc_labels else '?'} → {uc_labels[-1] if uc_labels else '?'})")
    else:
        print("  MDR: nenhum dado processado")

    # ── Exportar JSON ─────────────────────────────────────────────────────────
    import datetime
    output = {
        "gerado": datetime.date.today().isoformat(),
        "descricao": "Série histórica mensal backfill — DOTs e MDR HUSF",
        "ano_minimo": ANO_MINIMO,
        "dots": dots_merged,
        "mdrMensal": mdr_merged,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n  ✓ {OUTPUT_FILE} salvo.")
    print()

    # ── Próximo passo ─────────────────────────────────────────────────────────
    print("=" * 65)
    print("  Próximo passo:")
    print(f"  Coloque '{OUTPUT_FILE}' no mesmo diretório que data_iras.json")
    print("  e execute: Rscript husf_stats.R")
    print("  O script detectará automaticamente o arquivo e ativará")
    print("  os módulos CCF e CUSUM calibrado com série mensal completa.")
    print("=" * 65)


if __name__ == "__main__":
    main()
